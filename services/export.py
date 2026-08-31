import io
import json
import logging
from datetime import datetime, timezone

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    openpyxl = None

try:
    from lxml import etree
except ImportError:
    etree = None

from services.ingestion import VARIABLE_FIELDS

logger = logging.getLogger(__name__)

# Exact header order of the reference file's SDTM_LB/SDTM_VS worksheets
# (files/BC Examples.xlsx), so an exported specialization sheet is
# structurally identical to what the ingestion importer produces from a
# real CDISC-authored workbook — not just importer-parseable. The three
# *_version/vlm_source columns aren't tracked by DatasetSpecialization and
# are exported blank; VARIABLE_FIELDS (imported above) supplies the rest.
SPEC_SHEET_HEADER_FIELDS = [
    "package_date",
    "bc_id",
    "sdtmig_start_version",
    "sdtmig_end_version",
    "domain",
    "vlm_source",
    "vlm_group_id",
    "short_name",
]


BC_EXPORT_FIELDS = [
    "bc_id",
    "short_name",
    "definition",
    "ncit_code",
    "parent_bc_id",
    "bc_categories",
    "synonyms",
    "result_scales",
    "system",
    "system_name",
    "code",
    "package_date",
    "status",
]


def export_json(bc_list):
    """Export list of BC dicts to JSON string."""
    return json.dumps(bc_list, indent=2, default=str)


def export_xlsx(bc_list):
    """Export list of BC dicts to BytesIO XLSX."""
    if openpyxl is None:
        raise ImportError("openpyxl is required for XLSX export")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Biomedical Concepts"

    header_fill = PatternFill("solid", fgColor="003366")
    header_font_white = Font(bold=True, color="FFFFFF")

    for col_idx, field in enumerate(BC_EXPORT_FIELDS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field.replace("_", " ").title())
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, bc in enumerate(bc_list, start=2):
        for col_idx, field in enumerate(BC_EXPORT_FIELDS, start=1):
            value = bc.get("loinc_code", "") if field == "code" else bc.get(field, "")
            ws.cell(row=row_idx, column=col_idx, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


GOVERNANCE_BC_FIELDS = [
    "package_date",
    "short_name",
    "bc_id",
    "ncit_code",
    "parent_bc_id",
    "bc_categories",
    "synonyms",
    "result_scales",
    "definition",
    "system",
    "system_name",
    "code",
]
GOVERNANCE_DEC_FIELDS = ["dec_id", "ncit_dec_code", "dec_label", "data_type", "example_set"]
GOVERNANCE_HEADERS = GOVERNANCE_BC_FIELDS + GOVERNANCE_DEC_FIELDS + ["History of Change"]


def export_governance_xlsx(bc_objects, spec_objects=None):
    """Export BiomedicalConcept ORM objects in BC_LB worksheet format.

    Matches the column layout of the BC_LB sheet in the reference file:
    18 columns — BC fields, DEC fields, then History of Change.
    One BC-only row is written per BC, followed by one row per DEC
    (all BC fields repeated on each DEC row).

    When spec_objects (DatasetSpecialization ORM objects) is given, one
    "SDTM_<domain>" worksheet per domain is added — the same sheet naming
    and column headers (services.ingestion.SPEC_HEADER_FIELDS +
    VARIABLE_FIELDS) that services.ingestion.parse_xlsx expects, so the
    export can be re-imported through the ingestion pipeline unchanged.
    Returns a BytesIO object.
    """
    if openpyxl is None:
        raise ImportError("openpyxl is required for XLSX export")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BC_LB"

    header_fill = PatternFill("solid", fgColor="003366")
    header_font_white = Font(bold=True, color="FFFFFF")
    for col_idx, header in enumerate(GOVERNANCE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    for bc in bc_objects:
        bc_vals = {f: getattr(bc, f, "") or "" for f in GOVERNANCE_BC_FIELDS}
        bc_vals["code"] = bc.loinc_code or ""
        if not bc.loinc_code:
            bc_vals["system"] = ""
            bc_vals["system_name"] = ""
        bc_vals["History of Change"] = bc.history_of_change or ""

        # BC-only row (DEC columns left blank)
        for col_idx, header in enumerate(GOVERNANCE_HEADERS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=bc_vals.get(header, ""))
        row_idx += 1

        # One row per DEC, repeating BC fields
        from models.bc import DataElementConcept

        for dec in bc.decs.order_by(DataElementConcept.sort_order):
            for col_idx, header in enumerate(GOVERNANCE_HEADERS, start=1):
                if header in GOVERNANCE_BC_FIELDS:
                    val = bc_vals[header]
                elif header == "History of Change":
                    val = bc_vals["History of Change"]
                else:
                    val = getattr(dec, header, "") or ""
                ws.cell(row=row_idx, column=col_idx, value=val)
            row_idx += 1

    if spec_objects:
        spec_headers = SPEC_SHEET_HEADER_FIELDS + list(VARIABLE_FIELDS)

        specs_by_domain = {}
        for spec in spec_objects:
            specs_by_domain.setdefault(spec.domain or "", []).append(spec)

        for domain, specs in specs_by_domain.items():
            spec_ws = wb.create_sheet(f"SDTM_{domain}"[:31])
            for col_idx, header in enumerate(spec_headers, start=1):
                cell = spec_ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            row_idx = 2
            for spec in specs:
                # sdtmig_start_version/sdtmig_end_version/vlm_source are real
                # columns in the reference worksheet that this app's model
                # doesn't track — left blank rather than dropped, so the
                # exported header row matches the reference file exactly.
                header_vals = {
                    "vlm_group_id": spec.vlm_group_id,
                    "bc_id": spec.bc_id,
                    "domain": spec.domain or "",
                    "short_name": spec.short_name or "",
                    "package_date": "",
                    "sdtmig_start_version": "",
                    "sdtmig_end_version": "",
                    "vlm_source": "",
                }
                variables = spec.variables or [{}]
                for variable in variables:
                    for col_idx, header in enumerate(spec_headers, start=1):
                        value = header_vals[header] if header in header_vals else variable.get(header, "")
                        spec_ws.cell(row=row_idx, column=col_idx, value=value)
                    row_idx += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_odm_xml(bc_list):
    """Export BCs as ODM-XML string."""
    if etree is None:
        return "<ODM><!-- lxml not available --></ODM>"

    root = etree.Element(
        "ODM",
        attrib={
            "xmlns": "http://www.cdisc.org/ns/odm/v1.3",
            "FileType": "Snapshot",
            "FileOID": f'CDISC.BC.Export.{datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")}',
            "CreationDateTime": datetime.now(timezone.utc).isoformat(),
        },
    )

    for bc in bc_list:
        item_def = etree.SubElement(
            root,
            "ItemDef",
            attrib={
                "OID": bc.get("bc_id", "UNKNOWN"),
                "Name": bc.get("short_name", ""),
                "DataType": "text",
            },
        )
        desc = etree.SubElement(item_def, "Description")
        XML_NS = "http://www.w3.org/XML/1998/namespace"
        translated = etree.SubElement(desc, "TranslatedText", attrib={f"{{{XML_NS}}}lang": "en"})
        translated.text = bc.get("definition", "")
        if bc.get("ncit_code"):
            etree.SubElement(
                item_def,
                "Alias",
                attrib={
                    "Context": "nci:ExtCodeID",
                    "Name": bc.get("ncit_code", ""),
                },
            )

    return etree.tostring(root, pretty_print=True, xml_declaration=True, encoding="UTF-8").decode()
