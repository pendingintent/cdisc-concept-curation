"""Tests for routes/governance.py — Kanban advance and reject."""

import pytest
from sqlalchemy.exc import IntegrityError

from extensions import db
from models.audit import AuditLog
from models.bc import BiomedicalConcept
from models.governance import GovernanceRecord
from models.specialization import DatasetSpecialization

STATUS_ORDER = ["provisional", "sme_review", "cdisc_approval", "published"]


class TestGovernanceBoard:
    def test_board_returns_200(self, client):
        r = client.get("/governance/board")
        assert r.status_code == 200


class TestAdvance:
    def test_advances_provisional_to_sme_review(self, client, app, sample_bc):
        client.post("/governance/advance/C12345")
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            assert bc.status == "sme_review"

    def test_advance_through_all_stages(self, client, app, sample_bc):
        for expected in ["sme_review", "cdisc_approval", "published"]:
            client.post("/governance/advance/C12345")
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            assert bc.status == "published"

    def test_already_published_stays_published(self, client, app, sample_bc):
        # Advance to published
        for _ in range(3):
            client.post("/governance/advance/C12345")
        # Extra advance should not error or change status
        r = client.post("/governance/advance/C12345", follow_redirects=True)
        assert r.status_code == 200
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            assert bc.status == "published"

    def test_advance_creates_governance_record(self, client, app, sample_bc):
        client.post("/governance/advance/C12345")
        with app.app_context():
            rec = GovernanceRecord.query.filter_by(bc_id="C12345", action="advanced").first()
            assert rec is not None

    def test_advance_writes_audit_log(self, client, app, sample_bc):
        client.post("/governance/advance/C12345")
        with app.app_context():
            log = AuditLog.query.filter_by(entity_id="C12345", action="status_changed").first()
            assert log is not None
            assert log.before_state == {"status": "provisional"}
            assert log.after_state == {"status": "sme_review"}

    def test_advance_nonexistent_bc_returns_404(self, client):
        r = client.post("/governance/advance/NOPE")
        assert r.status_code == 404

    def test_advance_ajax_returns_json(self, client, sample_bc):
        r = client.post(
            "/governance/advance/C12345",
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        assert r.status_code == 200
        data = r.get_json()
        assert data["status"] == "sme_review"
        assert data["bc_id"] == "C12345"


class TestReject:
    def test_reject_returns_to_provisional(self, client, app, sample_bc):
        # First advance to sme_review, then reject
        client.post("/governance/advance/C12345")
        client.post("/governance/reject/C12345")
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            assert bc.status == "provisional"

    def test_reject_creates_governance_record(self, client, app, sample_bc):
        client.post("/governance/reject/C12345")
        with app.app_context():
            rec = GovernanceRecord.query.filter_by(bc_id="C12345", action="rejected").first()
            assert rec is not None

    def test_reject_writes_audit_log(self, client, app, sample_bc):
        client.post("/governance/advance/C12345")  # move to sme_review
        client.post("/governance/reject/C12345")
        with app.app_context():
            log = AuditLog.query.filter_by(entity_id="C12345", action="rejected").first()
            assert log is not None
            assert log.after_state == {"status": "provisional"}

    def test_reject_ajax_returns_json(self, client, sample_bc):
        r = client.post(
            "/governance/reject/C12345",
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        assert r.status_code == 200
        data = r.get_json()
        assert data["status"] == "provisional"

    def test_reject_nonexistent_bc_returns_404(self, client):
        r = client.post("/governance/reject/NOPE")
        assert r.status_code == 404

    def test_reject_from_published_returns_to_provisional(self, client, app, sample_bc):
        for _ in range(3):
            client.post("/governance/advance/C12345")
        client.post("/governance/reject/C12345")
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            assert bc.status == "provisional"

    def test_board_shows_reject_button_for_published_bc(self, client, app, sample_bc):
        for _ in range(3):
            client.post("/governance/advance/C12345")
        r = client.get("/governance/board")
        assert b'data-bc-id="C12345"' in r.data
        assert b"kanban-reject-btn" in r.data


class TestGovernanceExport:
    def test_export_returns_xlsx(self, client, app, sample_bc):
        for _ in range(3):
            client.post("/governance/advance/C12345")
        r = client.get("/governance/export")
        assert r.status_code == 200
        assert "spreadsheetml" in r.content_type

    def test_export_filename_in_content_disposition(self, client, app, sample_bc):
        for _ in range(3):
            client.post("/governance/advance/C12345")
        r = client.get("/governance/export?filename=my_report")
        assert "my_report.xlsx" in r.headers["Content-Disposition"]

    def test_export_enforces_xlsx_extension(self, client, app, sample_bc):
        for _ in range(3):
            client.post("/governance/advance/C12345")
        r = client.get("/governance/export?filename=my_report.csv")
        assert "my_report.xlsx" in r.headers["Content-Disposition"]

    def test_export_excludes_non_stage3_bcs(self, client, app, sample_bc):
        # BC stays provisional — no stage-3 governance record
        r = client.get("/governance/export")
        assert r.status_code == 200
        import io

        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(r.data))
        ws = wb.active
        assert ws.max_row == 1  # header row only

    def test_export_includes_stage3_bcs(self, client, app, sample_bc):
        for _ in range(3):
            client.post("/governance/advance/C12345")
        r = client.get("/governance/export")
        import io

        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(r.data))
        ws = wb.active
        assert ws.max_row >= 2  # at least one BC data row

    def test_export_system_columns_blank_without_loinc_code(self, client, app):
        with app.app_context():
            from extensions import db as _db

            bc = BiomedicalConcept(
                bc_id="C99998",
                short_name="No LOINC Concept",
                ncit_code="C99998",
                system="http://loinc.org/",
                system_name="LOINC",
                status="provisional",
            )
            _db.session.add(bc)
            _db.session.commit()
        for _ in range(3):
            client.post("/governance/advance/C99998")
        r = client.get("/governance/export")
        import io

        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(r.data))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        system_col = headers.index("system") + 1
        system_name_col = headers.index("system_name") + 1
        assert ws.cell(row=2, column=system_col).value in (None, "")
        assert ws.cell(row=2, column=system_name_col).value in (None, "")

    def test_export_system_columns_populated_with_loinc_code(self, client, app):
        with app.app_context():
            from extensions import db as _db

            bc = BiomedicalConcept(
                bc_id="C99997",
                short_name="LOINC System Concept",
                ncit_code="C99997",
                loinc_code="12345-6",
                system="http://loinc.org/",
                system_name="LOINC",
                status="provisional",
            )
            _db.session.add(bc)
            _db.session.commit()
        for _ in range(3):
            client.post("/governance/advance/C99997")
        r = client.get("/governance/export")
        import io

        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(r.data))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        system_col = headers.index("system") + 1
        system_name_col = headers.index("system_name") + 1
        assert ws.cell(row=2, column=system_col).value == "http://loinc.org/"
        assert ws.cell(row=2, column=system_name_col).value == "LOINC"

    def test_export_code_column_uses_loinc_code(self, client, app):
        with app.app_context():
            from extensions import db as _db

            bc = BiomedicalConcept(
                bc_id="C99999",
                short_name="LOINC Test Concept",
                ncit_code="C99999",
                loinc_code="12345-6",
                status="provisional",
            )
            _db.session.add(bc)
            _db.session.commit()
        for _ in range(3):
            client.post("/governance/advance/C99999")
        r = client.get("/governance/export")
        import io

        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(r.data))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        code_col = headers.index("code") + 1
        data_row = ws.cell(row=2, column=code_col).value
        assert data_row == "12345-6"


class TestSpecAdvance:
    def test_advances_provisional_to_sme_review(self, client, app, sample_spec):
        client.post(f"/governance/spec/advance/{sample_spec}")
        with app.app_context():
            spec = db.session.get(DatasetSpecialization, sample_spec)
            assert spec.status == "sme_review"

    def test_advance_through_all_stages(self, client, app, sample_spec):
        for _ in range(3):
            client.post(f"/governance/spec/advance/{sample_spec}")
        with app.app_context():
            spec = db.session.get(DatasetSpecialization, sample_spec)
            assert spec.status == "published"

    def test_already_published_stays_published(self, client, app, sample_spec):
        for _ in range(3):
            client.post(f"/governance/spec/advance/{sample_spec}")
        r = client.post(f"/governance/spec/advance/{sample_spec}", follow_redirects=True)
        assert r.status_code == 200
        with app.app_context():
            spec = db.session.get(DatasetSpecialization, sample_spec)
            assert spec.status == "published"

    def test_advance_creates_governance_record(self, client, app, sample_spec):
        client.post(f"/governance/spec/advance/{sample_spec}")
        with app.app_context():
            rec = GovernanceRecord.query.filter_by(vlm_group_id=sample_spec, action="advanced").first()
            assert rec is not None
            assert rec.bc_id is None

    def test_advance_writes_audit_log(self, client, app, sample_spec):
        client.post(f"/governance/spec/advance/{sample_spec}")
        with app.app_context():
            log = AuditLog.query.filter_by(entity_id=sample_spec, action="status_changed").first()
            assert log is not None
            assert log.before_state == {"status": "provisional"}
            assert log.after_state == {"status": "sme_review"}

    def test_advance_nonexistent_spec_returns_404(self, client):
        r = client.post("/governance/spec/advance/NOPE")
        assert r.status_code == 404

    def test_advance_ajax_returns_json(self, client, sample_spec):
        r = client.post(
            f"/governance/spec/advance/{sample_spec}",
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        assert r.status_code == 200
        data = r.get_json()
        assert data["status"] == "sme_review"
        assert data["vlm_group_id"] == sample_spec


class TestSpecReject:
    def test_reject_returns_to_provisional(self, client, app, sample_spec):
        client.post(f"/governance/spec/advance/{sample_spec}")
        client.post(f"/governance/spec/reject/{sample_spec}")
        with app.app_context():
            spec = db.session.get(DatasetSpecialization, sample_spec)
            assert spec.status == "provisional"

    def test_reject_creates_governance_record(self, client, app, sample_spec):
        client.post(f"/governance/spec/reject/{sample_spec}")
        with app.app_context():
            rec = GovernanceRecord.query.filter_by(vlm_group_id=sample_spec, action="rejected").first()
            assert rec is not None

    def test_reject_writes_audit_log(self, client, app, sample_spec):
        client.post(f"/governance/spec/advance/{sample_spec}")
        client.post(f"/governance/spec/reject/{sample_spec}")
        with app.app_context():
            log = AuditLog.query.filter_by(entity_id=sample_spec, action="rejected").first()
            assert log is not None
            assert log.after_state == {"status": "provisional"}

    def test_reject_ajax_returns_json(self, client, sample_spec):
        r = client.post(
            f"/governance/spec/reject/{sample_spec}",
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        assert r.status_code == 200
        data = r.get_json()
        assert data["status"] == "provisional"

    def test_reject_from_published_returns_to_provisional(self, client, app, sample_spec):
        for _ in range(3):
            client.post(f"/governance/spec/advance/{sample_spec}")
        client.post(f"/governance/spec/reject/{sample_spec}")
        with app.app_context():
            spec = db.session.get(DatasetSpecialization, sample_spec)
            assert spec.status == "provisional"

    def test_board_shows_reject_button_for_published_spec(self, client, app, sample_spec):
        for _ in range(3):
            client.post(f"/governance/spec/advance/{sample_spec}")
        r = client.get("/governance/board")
        assert f'data-vlm-group-id="{sample_spec}"'.encode() in r.data
        assert b"kanban-reject-btn" in r.data

    def test_reject_nonexistent_spec_returns_404(self, client):
        r = client.post("/governance/spec/reject/NOPE")
        assert r.status_code == 404


class TestGovernanceRecordOneEntityConstraint:
    def test_both_bc_id_and_vlm_group_id_set_raises(self, app, sample_spec):
        with app.app_context():
            db.session.add(GovernanceRecord(bc_id="C12345", vlm_group_id=sample_spec, stage=0, action="advanced"))
            with pytest.raises(IntegrityError):
                db.session.commit()
            db.session.rollback()

    def test_neither_bc_id_nor_vlm_group_id_set_raises(self, app):
        with app.app_context():
            db.session.add(GovernanceRecord(stage=0, action="advanced"))
            with pytest.raises(IntegrityError):
                db.session.commit()
            db.session.rollback()
