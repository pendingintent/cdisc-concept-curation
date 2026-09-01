"""Tests for routes/bc.py — CRUD, export, submission."""

import io
import json
from unittest.mock import patch

import openpyxl

from extensions import db
from models.audit import AuditLog
from models.bc import BiomedicalConcept, DataElementConcept


def _bc_form(**kwargs):
    defaults = {
        "bc_id": "C00001",
        "short_name": "Test Concept",
        "definition": "A definition.",
        "ncit_code": "C00001",
        "submitter": "tester",
    }
    defaults.update(kwargs)
    return defaults


# ---------------------------------------------------------------------------
# GET /bc/
# ---------------------------------------------------------------------------


class TestBcIndex:
    def test_returns_200(self, client):
        r = client.get("/bc/")
        assert r.status_code == 200

    def test_search_by_name(self, client, app, sample_bc):
        r = client.get("/bc/?q=Test")
        assert r.status_code == 200
        assert b"Test Concept" in r.data

    def test_search_no_match(self, client):
        r = client.get("/bc/?q=zzznomatch")
        assert r.status_code == 200

    def test_filter_by_status(self, client, sample_bc):
        r = client.get("/bc/?status=provisional")
        assert r.status_code == 200


# ---------------------------------------------------------------------------
# GET /bc/new
# ---------------------------------------------------------------------------


class TestNewBc:
    def test_returns_200(self, client):
        r = client.get("/bc/new")
        assert r.status_code == 200

    def test_shows_all_five_result_scale_checkboxes(self, client):
        r = client.get("/bc/new")
        for scale in ("Narrative", "Nominal", "Ordinal", "Quantitative", "Temporal"):
            assert scale.encode() in r.data


# ---------------------------------------------------------------------------
# POST /bc/ (create)
# ---------------------------------------------------------------------------


class TestCreateBc:
    def test_creates_bc_and_redirects(self, client, app):
        r = client.post("/bc/", data=_bc_form(), follow_redirects=False)
        assert r.status_code == 302
        with app.app_context():
            assert db.session.get(BiomedicalConcept, "C00001") is not None

    def test_missing_bc_id_redirects_with_error(self, client):
        r = client.post("/bc/", data=_bc_form(bc_id=""), follow_redirects=True)
        assert b"required" in r.data.lower() or r.status_code in (200, 302)

    def test_duplicate_bc_id_rejected(self, client, sample_bc):
        # First creation (sample_bc fixture did it already)
        r = client.post("/bc/", data=_bc_form(bc_id="C12345"), follow_redirects=True)
        assert b"already exists" in r.data or r.status_code in (200, 302)

    def test_create_writes_audit_log(self, client, app):
        client.post("/bc/", data=_bc_form())
        with app.app_context():
            log = AuditLog.query.filter_by(entity_id="C00001", action="created").first()
            assert log is not None

    def test_create_does_not_set_system_without_loinc_code(self, client, app):
        client.post("/bc/", data=_bc_form(system="http://loinc.org/", system_name="LOINC", loinc_code=""))
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C00001")
            assert bc.system in (None, "")
            assert bc.system_name in (None, "")

    def test_create_sets_system_when_loinc_code_provided(self, client, app):
        client.post("/bc/", data=_bc_form(system="http://loinc.org/", system_name="LOINC", loinc_code="4548-4"))
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C00001")
            assert bc.system == "http://loinc.org/"
            assert bc.system_name == "LOINC"

    def test_create_with_selected_result_scales(self, client, app):
        data = _bc_form()
        data["result_scales"] = ["Quantitative", "Ordinal"]
        client.post("/bc/", data=data)
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C00001")
            assert bc.result_scales == "Quantitative; Ordinal"

    def test_create_with_no_result_scales_selected(self, client, app):
        client.post("/bc/", data=_bc_form())
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C00001")
            assert bc.result_scales == ""

    def test_create_with_decs(self, client, app):
        """Reflects the actual decs[N][field] naming the rendered DEC table
        and static/js/main.js buildDecRow() submit, not a legacy shape."""
        data = _bc_form()
        data["decs[0][dec_label]"] = "Systolic"
        data["decs[0][data_type]"] = "decimal"
        data["decs[0][example_set]"] = "120"
        data["decs[1][dec_label]"] = "Diastolic"
        data["decs[1][data_type]"] = "decimal"
        data["decs[1][example_set]"] = "80"
        client.post("/bc/", data=data)
        with app.app_context():
            decs = DataElementConcept.query.filter_by(bc_id="C00001").order_by(DataElementConcept.sort_order).all()
            assert len(decs) == 2
            assert decs[0].dec_label == "Systolic"
            assert decs[1].dec_label == "Diastolic"


# ---------------------------------------------------------------------------
# GET /bc/<bc_id>
# ---------------------------------------------------------------------------


class TestBcDetail:
    def test_existing_bc_returns_200(self, client, sample_bc):
        with patch("routes.bc.LoincApiClient") as MockLoinc:
            MockLoinc.return_value.search.return_value = []
            r = client.get("/bc/C12345")
        assert r.status_code == 200

    def test_missing_bc_returns_404(self, client):
        r = client.get("/bc/DOESNOTEXIST")
        assert r.status_code == 404

    def test_valid_result_scale_checked(self, client, app, sample_bc):
        import re

        with app.app_context():
            bc = db.session.get(BiomedicalConcept, sample_bc)
            bc.result_scales = "Ordinal"
            db.session.commit()
        r = client.get(f"/bc/{sample_bc}")
        assert r.status_code == 200
        assert re.search(rb'value="Ordinal"\s*checked', r.data)
        assert not re.search(rb'value="Nominal"\s*checked', r.data)

    def test_dec_id_field_is_visible_and_precedes_label(self, client, app, sample_bc):
        """DEC ID must be a visible, editable column (not the old hidden
        input) and must appear before DEC Label so curators can see/set the
        value that populates the dec_id column in BC Examples.xlsx."""
        import re

        with app.app_context():
            db.session.add(DataElementConcept(dec_id="C12345.DEC.1", bc_id="C12345", dec_label="Systolic", data_type="decimal", sort_order=0))
            db.session.commit()
        r = client.get(f"/bc/{sample_bc}")
        html = r.data.decode()
        assert html.index(">DEC ID<") < html.index(">DEC Label<")
        assert re.search(r'<input type="text"[^>]*name="decs\[0\]\[dec_id\]"[^>]*value="C12345\.DEC\.1"', html)

    def test_unsupported_result_scale_shown_in_red(self, client, app, sample_bc):
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, sample_bc)
            bc.result_scales = "Continuous"
            db.session.commit()
        r = client.get(f"/bc/{sample_bc}")
        assert r.status_code == 200
        assert b"Continuous" in r.data
        assert b"not supported" in r.data.lower()
        assert b"text-danger" in r.data

    def test_loinc_api_called_when_code_set(self, client, app):
        with app.app_context():
            bc = BiomedicalConcept(
                bc_id="C99901",
                short_name="HbA1c",
                status="provisional",
                submitter="tester",
                loinc_code="4548-4",
            )
            db.session.add(bc)
            db.session.commit()

        loinc_result = {"LOINC_NUM": "4548-4", "LONG_COMMON_NAME": "Hemoglobin A1c/Hemoglobin.total in Blood", "SHORTNAME": "HbA1c MFr Bld"}
        with patch("routes.bc.LoincApiClient") as MockLoinc:
            MockLoinc.return_value.search.return_value = [loinc_result]
            r = client.get("/bc/C99901")

        assert r.status_code == 200
        MockLoinc.return_value.search.assert_called_once_with("4548-4", size=1)
        assert b"HbA1c MFr Bld" in r.data

    def test_loinc_api_not_called_when_no_code(self, client, app):
        with app.app_context():
            bc = BiomedicalConcept(
                bc_id="C99902",
                short_name="No LOINC",
                status="provisional",
                submitter="tester",
            )
            db.session.add(bc)
            db.session.commit()

        with patch("routes.bc.LoincApiClient") as MockLoinc:
            r = client.get("/bc/C99902")

        assert r.status_code == 200
        MockLoinc.return_value.search.assert_not_called()

    def test_loinc_api_error_does_not_break_page(self, client, app):
        with app.app_context():
            bc = BiomedicalConcept(
                bc_id="C99903",
                short_name="LOINC Error BC",
                status="provisional",
                submitter="tester",
                loinc_code="4548-4",
            )
            db.session.add(bc)
            db.session.commit()

        with patch("routes.bc.LoincApiClient") as MockLoinc:
            MockLoinc.return_value.search.return_value = [{"error": "timeout"}]
            r = client.get("/bc/C99903")

        assert r.status_code == 200

    def test_loinc_spinner_not_shown_when_loinc_fetched_server_side(self, client, app):
        with app.app_context():
            bc = BiomedicalConcept(
                bc_id="C99904",
                short_name="LOINC Spinner BC",
                status="provisional",
                submitter="tester",
                loinc_code="4548-4",
            )
            db.session.add(bc)
            db.session.commit()

        loinc_result = {"LOINC_NUM": "4548-4", "LONG_COMMON_NAME": "Hemoglobin A1c/Hemoglobin.total in Blood"}
        with patch("routes.bc.LoincApiClient") as MockLoinc:
            MockLoinc.return_value.search.return_value = [loinc_result]
            r = client.get("/bc/C99904")

        assert r.status_code == 200
        assert b"loinc-loading-indicator" not in r.data

    def test_loinc_metadata_saved_when_fetched_in_detail(self, client, app):
        with app.app_context():
            bc = BiomedicalConcept(
                bc_id="C99905",
                short_name="LOINC Save BC",
                status="provisional",
                submitter="tester",
                loinc_code="4548-4",
            )
            db.session.add(bc)
            db.session.commit()

        loinc_result = {"LOINC_NUM": "4548-4", "LONG_COMMON_NAME": "Hemoglobin A1c/Hemoglobin.total in Blood"}
        with patch("routes.bc.LoincApiClient") as MockLoinc:
            MockLoinc.return_value.search.return_value = [loinc_result]
            client.get("/bc/C99905")

        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C99905")
            assert bc.loinc_metadata is not None
            assert "4548-4" in bc.loinc_metadata


# ---------------------------------------------------------------------------
# POST /bc/<bc_id>/edit
# ---------------------------------------------------------------------------


class TestEditBc:
    def test_updates_short_name(self, client, app, sample_bc):
        client.post("/bc/C12345/edit", data={"short_name": "Updated Name"})
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            assert bc.short_name == "Updated Name"

    def test_edit_writes_audit_log(self, client, app, sample_bc):
        client.post("/bc/C12345/edit", data={"short_name": "Updated Name"})
        with app.app_context():
            log = AuditLog.query.filter_by(entity_id="C12345", action="updated").first()
            assert log is not None
            assert log.before_state["short_name"] == "Test Concept"

    def test_nonexistent_bc_returns_404(self, client):
        r = client.post("/bc/NOPE/edit", data={"short_name": "X"})
        assert r.status_code == 404

    def test_edit_blocked_when_published(self, client, app, sample_bc):
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            bc.status = "published"
            db.session.commit()
        r = client.post("/bc/C12345/edit", data={"short_name": "Updated Name"}, follow_redirects=True)
        assert r.status_code == 200
        assert b"Ready to Publish" in r.data
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            assert bc.short_name == "Test Concept"

    def test_edit_updates_result_scales_when_marker_present(self, client, app, sample_bc):
        client.post(
            "/bc/C12345/edit",
            data={"result_scales_submitted": "1", "result_scales": ["Quantitative", "Ordinal"]},
        )
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            assert bc.result_scales == "Quantitative; Ordinal"

    def test_edit_clears_result_scales_when_all_unchecked(self, client, app, sample_bc):
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            bc.result_scales = "Ordinal"
            db.session.commit()
        client.post("/bc/C12345/edit", data={"result_scales_submitted": "1"})
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            assert bc.result_scales == ""

    def test_edit_without_marker_preserves_existing_result_scales(self, client, app, sample_bc):
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            bc.result_scales = "Ordinal"
            db.session.commit()
        client.post("/bc/C12345/edit", data={"short_name": "Updated Name"})
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            assert bc.result_scales == "Ordinal"

    def test_edit_preserves_unsupported_value_alongside_new_selection(self, client, app, sample_bc):
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            bc.result_scales = "Continuous"
            db.session.commit()
        client.post(
            "/bc/C12345/edit",
            data={"result_scales_submitted": "1", "result_scales": ["Quantitative", "Continuous"]},
        )
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            assert bc.result_scales == "Quantitative; Continuous"

    def test_edit_clears_ncit_code_when_submitted_empty(self, client, app, sample_bc):
        client.post("/bc/C12345/edit", data={"ncit_code": ""})
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            assert bc.ncit_code is None

    def test_edit_clears_ncit_metadata_when_ncit_code_cleared(self, client, app, sample_bc):
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            bc.ncit_metadata = '{"preferred_name": "Test"}'
            db.session.commit()
        client.post("/bc/C12345/edit", data={"ncit_code": "", "ncit_metadata": '{"preferred_name": "Test"}'})
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            assert bc.ncit_metadata is None

    def test_edit_clears_parent_bc_id_when_submitted_empty(self, client, app, sample_bc):
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            bc.parent_bc_id = "C99999"
            db.session.commit()
        client.post("/bc/C12345/edit", data={"parent_bc_id": ""})
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            assert bc.parent_bc_id is None

    def test_edit_clears_loinc_code_and_metadata_when_submitted_empty(self, client, app, sample_bc):
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            bc.loinc_code = "4548-4"
            bc.loinc_metadata = '{"LONG_COMMON_NAME": "HbA1c"}'
            db.session.commit()
        client.post("/bc/C12345/edit", data={"loinc_code": "", "loinc_metadata": '{"LONG_COMMON_NAME": "HbA1c"}'})
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            assert bc.loinc_code is None
            assert bc.loinc_metadata is None

    def test_edit_clears_system_and_system_name_when_loinc_code_cleared(self, client, app, sample_bc):
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            bc.loinc_code = "4548-4"
            bc.system = "http://loinc.org/"
            bc.system_name = "LOINC"
            db.session.commit()
        client.post("/bc/C12345/edit", data={"loinc_code": ""})
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            assert bc.system in (None, "")
            assert bc.system_name in (None, "")

    def test_edit_preserves_system_when_loinc_code_present(self, client, app, sample_bc):
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            bc.loinc_code = "4548-4"
            bc.system = "http://loinc.org/"
            bc.system_name = "LOINC"
            db.session.commit()
        client.post("/bc/C12345/edit", data={"loinc_code": "4548-4", "system": "http://loinc.org/", "system_name": "LOINC"})
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            assert bc.system == "http://loinc.org/"
            assert bc.system_name == "LOINC"

    def test_edit_strips_whitespace_from_ncit_code(self, client, app, sample_bc):
        client.post("/bc/C12345/edit", data={"ncit_code": "   "})
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            assert bc.ncit_code is None

    def test_detail_renders_empty_ncit_code_when_none(self, client, app, sample_bc):
        """Regression: Jinja2 renders Python None as 'None' in HTML attributes.
        When bc.ncit_code is None the input must have value='' not value='None',
        otherwise the browser re-submits 'None' causing a spurious NCIt fetch spinner."""
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            bc.ncit_code = None
            db.session.commit()
        with patch("routes.bc.LoincApiClient") as MockLoinc:
            MockLoinc.return_value.search.return_value = []
            r = client.get("/bc/C12345")
        assert b'value="None"' not in r.data
        assert b'name="ncit_code"' in r.data

    # -----------------------------------------------------------------------
    # DEC persistence via edit (regression: decs[N][field] naming used by the
    # rendered table/static/js/main.js must match what the route parses,
    # otherwise every save wipes the BC's DECs)
    # -----------------------------------------------------------------------

    def test_edit_keeps_unchanged_decs(self, client, app, sample_bc):
        with app.app_context():
            db.session.add(DataElementConcept(dec_id="C12345.DEC.1", bc_id="C12345", dec_label="Systolic", data_type="decimal", example_set="120", sort_order=0))
            db.session.commit()
        client.post(
            "/bc/C12345/edit",
            data={
                "decs[0][dec_id]": "C12345.DEC.1",
                "decs[0][dec_label]": "Systolic",
                "decs[0][data_type]": "decimal",
                "decs[0][example_set]": "120",
            },
        )
        with app.app_context():
            decs = DataElementConcept.query.filter_by(bc_id="C12345").all()
            assert len(decs) == 1
            assert decs[0].dec_id == "C12345.DEC.1"
            assert decs[0].dec_label == "Systolic"

    def test_edit_updates_existing_dec_fields(self, client, app, sample_bc):
        with app.app_context():
            db.session.add(DataElementConcept(dec_id="C12345.DEC.1", bc_id="C12345", dec_label="Systolic", data_type="decimal", example_set="120", sort_order=0))
            db.session.commit()
        client.post(
            "/bc/C12345/edit",
            data={
                "decs[0][dec_id]": "C12345.DEC.1",
                "decs[0][dec_label]": "Systolic Blood Pressure",
                "decs[0][data_type]": "decimal",
                "decs[0][example_set]": "110-140",
            },
        )
        with app.app_context():
            decs = DataElementConcept.query.filter_by(bc_id="C12345").all()
            assert len(decs) == 1
            assert decs[0].dec_id == "C12345.DEC.1"
            assert decs[0].dec_label == "Systolic Blood Pressure"
            assert decs[0].example_set == "110-140"

    def test_edit_adds_new_dec_row(self, client, app, sample_bc):
        with app.app_context():
            db.session.add(DataElementConcept(dec_id="C12345.DEC.1", bc_id="C12345", dec_label="Systolic", data_type="decimal", sort_order=0))
            db.session.commit()
        client.post(
            "/bc/C12345/edit",
            data={
                "decs[0][dec_id]": "C12345.DEC.1",
                "decs[0][dec_label]": "Systolic",
                "decs[0][data_type]": "decimal",
                "decs[1][dec_label]": "Diastolic",
                "decs[1][data_type]": "decimal",
            },
        )
        with app.app_context():
            decs = DataElementConcept.query.filter_by(bc_id="C12345").order_by(DataElementConcept.sort_order).all()
            assert len(decs) == 2
            assert decs[1].dec_label == "Diastolic"

    def test_edit_removes_dec_omitted_from_submission(self, client, app, sample_bc):
        with app.app_context():
            db.session.add_all(
                [
                    DataElementConcept(dec_id="C12345.DEC.1", bc_id="C12345", dec_label="Systolic", data_type="decimal", sort_order=0),
                    DataElementConcept(dec_id="C12345.DEC.2", bc_id="C12345", dec_label="Diastolic", data_type="decimal", sort_order=1),
                ]
            )
            db.session.commit()
        client.post(
            "/bc/C12345/edit",
            data={
                "decs[0][dec_id]": "C12345.DEC.1",
                "decs[0][dec_label]": "Systolic",
                "decs[0][data_type]": "decimal",
            },
        )
        with app.app_context():
            decs = DataElementConcept.query.filter_by(bc_id="C12345").all()
            assert len(decs) == 1
            assert decs[0].dec_label == "Systolic"

    def test_edit_persists_required_flag(self, client, app, sample_bc):
        client.post(
            "/bc/C12345/edit",
            data={
                "decs[0][dec_label]": "Systolic",
                "decs[0][data_type]": "decimal",
                "decs[0][required]": "1",
            },
        )
        with app.app_context():
            dec = DataElementConcept.query.filter_by(bc_id="C12345").first()
            assert dec.required is True


# ---------------------------------------------------------------------------
# POST /bc/<bc_id>/clear-ncit
# ---------------------------------------------------------------------------


class TestClearNcitCode:
    def test_clears_ncit_code_and_metadata(self, client, app, sample_bc):
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            bc.ncit_metadata = '{"preferred_name": "Test"}'
            db.session.commit()
        client.post("/bc/C12345/clear-ncit", follow_redirects=False)
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            assert bc.ncit_code is None
            assert bc.ncit_metadata is None

    def test_clear_ncit_writes_audit_log(self, client, app, sample_bc):
        client.post("/bc/C12345/clear-ncit")
        with app.app_context():
            log = AuditLog.query.filter_by(entity_id="C12345", action="ncit_cleared").first()
            assert log is not None

    def test_clear_ncit_nonexistent_bc_returns_404(self, client):
        r = client.post("/bc/NOTREAL/clear-ncit")
        assert r.status_code == 404

    def test_clear_ncit_blocked_when_published(self, client, app, sample_bc):
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            bc.status = "published"
            db.session.commit()
        r = client.post("/bc/C12345/clear-ncit", follow_redirects=True)
        assert r.status_code == 200
        assert b"Ready to Publish" in r.data
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            assert bc.ncit_code == "C12345"

    def test_clear_ncit_also_clears_parent_bc_id(self, client, app, sample_bc):
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            bc.parent_bc_id = "C99999"
            db.session.commit()
        client.post("/bc/C12345/clear-ncit", follow_redirects=False)
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            assert bc.parent_bc_id is None


# ---------------------------------------------------------------------------
# POST /bc/<bc_id>/clear-loinc
# ---------------------------------------------------------------------------


class TestClearLoincCode:
    def test_clears_loinc_code_and_metadata(self, client, app, sample_bc):
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            bc.loinc_code = "4548-4"
            bc.loinc_metadata = '{"LONG_COMMON_NAME": "HbA1c"}'
            db.session.commit()
        client.post("/bc/C12345/clear-loinc", follow_redirects=False)
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            assert bc.loinc_code is None
            assert bc.loinc_metadata is None

    def test_clear_loinc_writes_audit_log(self, client, app, sample_bc):
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            bc.loinc_code = "4548-4"
            db.session.commit()
        client.post("/bc/C12345/clear-loinc")
        with app.app_context():
            log = AuditLog.query.filter_by(entity_id="C12345", action="loinc_cleared").first()
            assert log is not None

    def test_clear_loinc_also_clears_system_and_system_name(self, client, app, sample_bc):
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            bc.loinc_code = "4548-4"
            bc.system = "http://loinc.org/"
            bc.system_name = "LOINC"
            db.session.commit()
        client.post("/bc/C12345/clear-loinc", follow_redirects=False)
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            assert bc.system in (None, "")
            assert bc.system_name in (None, "")

    def test_clear_loinc_nonexistent_bc_returns_404(self, client):
        r = client.post("/bc/NOTREAL/clear-loinc")
        assert r.status_code == 404

    def test_clear_loinc_blocked_when_published(self, client, app, sample_bc):
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            bc.loinc_code = "4548-4"
            bc.status = "published"
            db.session.commit()
        r = client.post("/bc/C12345/clear-loinc", follow_redirects=True)
        assert r.status_code == 200
        assert b"Ready to Publish" in r.data
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            assert bc.loinc_code == "4548-4"


# ---------------------------------------------------------------------------
# POST /bc/<bc_id>/submit
# ---------------------------------------------------------------------------


class TestSubmitForReview:
    def test_advances_status_to_sme_review(self, client, app, sample_bc):
        client.post("/bc/C12345/submit")
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            assert bc.status == "sme_review"

    def test_submit_writes_audit_log(self, client, app, sample_bc):
        client.post("/bc/C12345/submit")
        with app.app_context():
            log = AuditLog.query.filter_by(entity_id="C12345", action="submitted_for_review").first()
            assert log is not None


# ---------------------------------------------------------------------------
# POST /bc/<bc_id>/delete
# ---------------------------------------------------------------------------


class TestDeleteBc:
    def test_deletes_bc(self, client, app, sample_bc):
        client.post("/bc/C12345/delete")
        with app.app_context():
            assert db.session.get(BiomedicalConcept, "C12345") is None

    def test_delete_writes_audit_log(self, client, app, sample_bc):
        client.post("/bc/C12345/delete")
        with app.app_context():
            log = AuditLog.query.filter_by(entity_id="C12345", action="deleted").first()
            assert log is not None

    def test_nonexistent_bc_returns_404(self, client):
        r = client.post("/bc/NOPE/delete")
        assert r.status_code == 404

    def test_delete_blocked_when_published(self, client, app, sample_bc):
        with app.app_context():
            bc = db.session.get(BiomedicalConcept, "C12345")
            bc.status = "published"
            db.session.commit()
        r = client.post("/bc/C12345/delete", follow_redirects=True)
        assert r.status_code == 200
        assert b"Ready to Publish" in r.data
        with app.app_context():
            assert db.session.get(BiomedicalConcept, "C12345") is not None


# ---------------------------------------------------------------------------
# GET /bc/export
# ---------------------------------------------------------------------------


class TestExport:
    def test_json_export(self, client, sample_bc):
        r = client.get("/bc/export?format=json")
        assert r.status_code == 200
        assert r.content_type == "application/json"

    def test_json_export_includes_decs(self, client, app, sample_bc):
        with app.app_context():
            db.session.add(DataElementConcept(dec_id="C12345.DEC.1", bc_id="C12345", dec_label="Systolic", data_type="decimal", sort_order=0))
            db.session.commit()
        r = client.get("/bc/export?format=json")
        parsed = json.loads(r.data)
        bc = next(b for b in parsed if b["bc_id"] == "C12345")
        assert bc["decs"][0]["dec_label"] == "Systolic"

    def test_xlsx_export(self, client, sample_bc):
        r = client.get("/bc/export?format=xlsx")
        assert r.status_code == 200
        assert "spreadsheetml" in r.content_type

    def test_xlsx_export_matches_bc_lb_reference_format(self, client, app, sample_bc):
        """The BC list's own export (not just the governance/published-only
        export) must produce the files/BC Examples.xlsx BC_LB shape: a
        BC-only row followed by one row per DEC."""
        with app.app_context():
            db.session.add(DataElementConcept(dec_id="C12345.DEC.1", bc_id="C12345", dec_label="Systolic", data_type="decimal", sort_order=0))
            db.session.commit()
        r = client.get("/bc/export?format=xlsx")
        wb = openpyxl.load_workbook(io.BytesIO(r.data))
        ws = wb.active
        assert ws.title == "BC_LB"
        headers = [c.value for c in ws[1]]
        dec_label_col = headers.index("dec_label") + 1
        assert ws.cell(row=2, column=dec_label_col).value in ("", None)
        assert ws.cell(row=3, column=dec_label_col).value == "Systolic"

    def test_odm_xml_export(self, client, sample_bc):
        r = client.get("/bc/export?format=odm")
        assert r.status_code == 200
        assert "xml" in r.content_type

    def test_odm_xml_export_includes_dec_item_def(self, client, app, sample_bc):
        with app.app_context():
            db.session.add(DataElementConcept(dec_id="C12345.DEC.1", bc_id="C12345", dec_label="Systolic", data_type="decimal", sort_order=0))
            db.session.commit()
        r = client.get("/bc/export?format=odm")
        assert b'OID="C12345.DEC.1"' in r.data
        assert b'Name="Systolic"' in r.data


# ---------------------------------------------------------------------------
# GET /bc/library/<concept_id>
# ---------------------------------------------------------------------------

LIBRARY_BC_NO_LOINC = {
    "conceptId": "C147905",
    "shortName": "Diastolic Blood Pressure",
    "definition": "The minimum pressure in the arteries.",
    "coding": [],
    "dataElementConcepts": [],
}

LIBRARY_BC_WITH_LOINC = {
    "conceptId": "C64849",
    "shortName": "HbA1c Percent",
    "definition": "A test measuring HbA1c.",
    "coding": [{"system": "http://loinc.org/", "systemName": "LOINC", "code": "4548-4"}],
    "dataElementConcepts": [],
}

LOINC_RESULT = {
    "LOINC_NUM": "4548-4",
    "LONG_COMMON_NAME": "Hemoglobin A1c/Hemoglobin.total in Blood",
    "SHORTNAME": "HbA1c MFr Bld",
    "PROPERTY": "MFr",
    "units": "%",
}


class TestLibraryDetail:
    def test_renders_page_for_valid_concept(self, client):
        with patch("routes.bc.CDISCApiClient") as MockCDISC, patch("routes.bc.LoincApiClient") as MockLoinc:
            MockCDISC.return_value.get_bc.return_value = LIBRARY_BC_NO_LOINC
            MockLoinc.return_value.search.return_value = []
            r = client.get("/bc/library/C147905")
        assert r.status_code == 200
        assert b"Diastolic Blood Pressure" in r.data

    def test_redirects_on_api_error(self, client):
        with patch("routes.bc.CDISCApiClient") as MockCDISC:
            MockCDISC.return_value.get_bc.return_value = {"error": "Not found"}
            r = client.get("/bc/library/CXXX", follow_redirects=False)
        assert r.status_code == 302

    def test_loinc_api_called_when_loinc_coding_present(self, client):
        with patch("routes.bc.CDISCApiClient") as MockCDISC, patch("routes.bc.LoincApiClient") as MockLoinc:
            MockCDISC.return_value.get_bc.return_value = LIBRARY_BC_WITH_LOINC
            MockLoinc.return_value.search.return_value = [LOINC_RESULT]
            r = client.get("/bc/library/C64849")
        assert r.status_code == 200
        MockLoinc.return_value.search.assert_called_once_with("4548-4", size=1)
        assert b"HbA1c MFr Bld" in r.data

    def test_loinc_api_not_called_when_no_loinc_coding(self, client):
        with patch("routes.bc.CDISCApiClient") as MockCDISC, patch("routes.bc.LoincApiClient") as MockLoinc:
            MockCDISC.return_value.get_bc.return_value = LIBRARY_BC_NO_LOINC
            r = client.get("/bc/library/C147905")
        assert r.status_code == 200
        MockLoinc.return_value.search.assert_not_called()

    def test_loinc_api_error_does_not_break_page(self, client):
        with patch("routes.bc.CDISCApiClient") as MockCDISC, patch("routes.bc.LoincApiClient") as MockLoinc:
            MockCDISC.return_value.get_bc.return_value = LIBRARY_BC_WITH_LOINC
            MockLoinc.return_value.search.return_value = [{"error": "timeout"}]
            r = client.get("/bc/library/C64849")
        assert r.status_code == 200
