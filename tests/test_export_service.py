"""Tests for services/export.py — JSON, XLSX, governance XLSX, ODM-XML."""

import json

import openpyxl
from lxml import etree

from extensions import db
from models.bc import BiomedicalConcept, DataElementConcept
from models.specialization import DatasetSpecialization
from services.export import (
    GOVERNANCE_HEADERS,
    export_governance_xlsx,
    export_json,
    export_odm_xml,
)
from services.ingestion import VARIABLE_FIELDS, parse_xlsx

SAMPLE_BCS = [
    {
        "bc_id": "C64849",
        "short_name": "Hemoglobin A1c Measurement",
        "definition": "A quantitative measurement of HbA1c.",
        "ncit_code": "C64849",
        "parent_bc_id": None,
        "bc_categories": "Laboratory Tests",
        "synonyms": "HbA1c;A1C",
        "result_scales": "Quantitative",
        "system": "http://loinc.org/",
        "system_name": "LOINC",
        "loinc_code": "4548-4",
        "package_date": "2026-01-01",
        "status": "provisional",
        "submitter": "tester",
    },
    {
        "bc_id": "C25298",
        "short_name": "Systolic Blood Pressure",
        "definition": "The maximum arterial pressure.",
        "ncit_code": "",
        "parent_bc_id": None,
        "bc_categories": "Vital Signs",
        "synonyms": "",
        "result_scales": "Quantitative",
        "system": "",
        "system_name": "",
        "loinc_code": "",
        "package_date": "",
        "status": "sme_review",
        "submitter": "tester",
    },
]


class TestExportJson:
    def test_round_trip(self):
        out = export_json(SAMPLE_BCS)
        parsed = json.loads(out)
        assert len(parsed) == 2
        assert parsed[0]["bc_id"] == "C64849"
        assert parsed[1]["short_name"] == "Systolic Blood Pressure"

    def test_empty_list(self):
        assert json.loads(export_json([])) == []

    def test_non_serializable_values_coerced(self):
        from datetime import datetime

        out = export_json([{"bc_id": "X", "created_at": datetime(2026, 1, 1)}])
        assert "2026-01-01" in out


class TestExportGovernanceXlsx:
    def _make_bc_with_decs(self):
        bc = BiomedicalConcept(
            bc_id="C64849",
            short_name="Hemoglobin A1c Measurement",
            definition="A quantitative measurement of HbA1c.",
            ncit_code="C64849",
            loinc_code="4548-4",
            system="http://loinc.org/",
            system_name="LOINC",
            history_of_change="Initial version",
            status="provisional",
            submitter="tester",
        )
        db.session.add(bc)
        db.session.add_all(
            [
                DataElementConcept(
                    dec_id="C64849.DEC.1",
                    bc_id="C64849",
                    dec_label="Result Value",
                    data_type="decimal",
                    sort_order=0,
                ),
                DataElementConcept(
                    dec_id="C64849.DEC.2",
                    bc_id="C64849",
                    dec_label="Unit",
                    data_type="string",
                    sort_order=1,
                ),
            ]
        )
        db.session.commit()
        return bc

    def test_bc_row_then_dec_rows(self, app):
        with app.app_context():
            bc = self._make_bc_with_decs()
            wb = openpyxl.load_workbook(export_governance_xlsx([bc]))
        ws = wb.active
        assert ws.title == "BC_LB"
        headers = [c.value for c in ws[1]]
        assert headers == GOVERNANCE_HEADERS
        # Row 2: BC-only row — DEC columns blank, loinc code in "code" col
        code_col = GOVERNANCE_HEADERS.index("code") + 1
        dec_label_col = GOVERNANCE_HEADERS.index("dec_label") + 1
        assert ws.cell(row=2, column=code_col).value == "4548-4"
        assert ws.cell(row=2, column=dec_label_col).value in ("", None)
        # Rows 3-4: one per DEC in sort order, BC fields repeated
        assert ws.cell(row=3, column=dec_label_col).value == "Result Value"
        assert ws.cell(row=4, column=dec_label_col).value == "Unit"
        bc_id_col = GOVERNANCE_HEADERS.index("bc_id") + 1
        assert ws.cell(row=3, column=bc_id_col).value == "C64849"
        # History of Change lands in the last column
        assert ws.cell(row=2, column=len(GOVERNANCE_HEADERS)).value == "Initial version"

    def test_ncit_dec_code_column_mirrors_dec_id(self, app):
        """Column N (ncit_dec_code) must always mirror column M (dec_id) on
        export — the DEC ID field is the only identifier curators can set
        through the UI, so the legacy ncit_dec_code column is populated from
        it rather than the (now unused) stored value."""
        with app.app_context():
            bc = self._make_bc_with_decs()
            wb = openpyxl.load_workbook(export_governance_xlsx([bc]))
        ws = wb.active
        dec_id_col = GOVERNANCE_HEADERS.index("dec_id") + 1
        ncit_dec_code_col = GOVERNANCE_HEADERS.index("ncit_dec_code") + 1
        for row in (3, 4):
            assert ws.cell(row=row, column=ncit_dec_code_col).value == ws.cell(row=row, column=dec_id_col).value

    def test_ncit_dec_code_column_overrides_a_stored_distinct_value(self, app):
        with app.app_context():
            bc = BiomedicalConcept(bc_id="C111", short_name="X", status="provisional")
            db.session.add(bc)
            db.session.add(
                DataElementConcept(
                    dec_id="C111.DEC.1",
                    bc_id="C111",
                    ncit_dec_code="SOME_OTHER_CODE",
                    dec_label="Label",
                    data_type="string",
                    sort_order=0,
                )
            )
            db.session.commit()
            wb = openpyxl.load_workbook(export_governance_xlsx([bc]))
        ws = wb.active
        dec_id_col = GOVERNANCE_HEADERS.index("dec_id") + 1
        ncit_dec_code_col = GOVERNANCE_HEADERS.index("ncit_dec_code") + 1
        assert ws.cell(row=3, column=dec_id_col).value == "C111.DEC.1"
        assert ws.cell(row=3, column=ncit_dec_code_col).value == "C111.DEC.1"

    def test_bc_without_loinc_blanks_system_fields(self, app):
        with app.app_context():
            bc = BiomedicalConcept(
                bc_id="C25298",
                short_name="Systolic Blood Pressure",
                system="http://should-be-blanked/",
                system_name="STALE",
                status="provisional",
            )
            db.session.add(bc)
            db.session.commit()
            wb = openpyxl.load_workbook(export_governance_xlsx([bc]))
        ws = wb.active
        system_col = GOVERNANCE_HEADERS.index("system") + 1
        system_name_col = GOVERNANCE_HEADERS.index("system_name") + 1
        assert ws.cell(row=2, column=system_col).value in ("", None)
        assert ws.cell(row=2, column=system_name_col).value in ("", None)


class TestExportGovernanceXlsxSpecializations:
    def _make_bc(self, bc_id="C64854"):
        bc = BiomedicalConcept(bc_id=bc_id, short_name="Ketone Concentration in Urine", status="published")
        db.session.add(bc)
        db.session.commit()
        return bc

    def test_no_specializations_produces_no_extra_sheets(self, app):
        with app.app_context():
            self._make_bc()
            wb = openpyxl.load_workbook(export_governance_xlsx([], []))
        assert wb.sheetnames == ["BC_LB"]

    def test_specializations_grouped_by_domain_into_sdtm_sheets(self, app):
        with app.app_context():
            bc = self._make_bc()
            db.session.add_all(
                [
                    DatasetSpecialization(vlm_group_id="C64854.LB", bc_id=bc.bc_id, domain="LB", short_name="Ketones LB", status="published"),
                    DatasetSpecialization(vlm_group_id="C64854.VS", bc_id=bc.bc_id, domain="VS", short_name="Ketones VS", status="published"),
                ]
            )
            db.session.commit()
            specs = DatasetSpecialization.query.order_by(DatasetSpecialization.vlm_group_id).all()
            wb = openpyxl.load_workbook(export_governance_xlsx([], specs))
        assert "SDTM_LB" in wb.sheetnames
        assert "SDTM_VS" in wb.sheetnames

    def test_sheet_headers_match_import_expectations(self, app):
        with app.app_context():
            bc = self._make_bc()
            spec = DatasetSpecialization(vlm_group_id="C64854.LB", bc_id=bc.bc_id, domain="LB", short_name="Ketones LB", status="published")
            db.session.add(spec)
            db.session.commit()
            wb = openpyxl.load_workbook(export_governance_xlsx([], [spec]))
        ws = wb["SDTM_LB"]
        headers = [c.value for c in ws[1]]
        # Exact header row from the reference file's SDTM_LB/SDTM_VS sheets
        # (files/BC Examples.xlsx) — order matters here, not just column
        # membership, since the goal is a byte-for-byte-comparable header.
        assert headers == [
            "package_date",
            "bc_id",
            "sdtmig_start_version",
            "sdtmig_end_version",
            "domain",
            "vlm_source",
            "vlm_group_id",
            "short_name",
        ] + list(VARIABLE_FIELDS)

    def test_one_row_per_variable_with_header_fields_repeated(self, app):
        with app.app_context():
            bc = self._make_bc()
            spec = DatasetSpecialization(vlm_group_id="C64854.LB", bc_id=bc.bc_id, domain="LB", short_name="Ketones LB", status="published")
            spec.variables = [
                {"sdtm_variable": "LBTESTCD", "data_type": "string"},
                {"sdtm_variable": "LBORRES", "data_type": "decimal"},
            ]
            db.session.add(spec)
            db.session.commit()
            wb = openpyxl.load_workbook(export_governance_xlsx([], [spec]))
        ws = wb["SDTM_LB"]
        headers = [c.value for c in ws[1]]
        vlm_col = headers.index("vlm_group_id") + 1
        var_col = headers.index("sdtm_variable") + 1
        assert ws.cell(row=2, column=vlm_col).value == "C64854.LB"
        assert ws.cell(row=2, column=var_col).value == "LBTESTCD"
        assert ws.cell(row=3, column=vlm_col).value == "C64854.LB"
        assert ws.cell(row=3, column=var_col).value == "LBORRES"

    def test_specialization_with_no_variables_still_exports_header_row(self, app):
        with app.app_context():
            bc = self._make_bc()
            spec = DatasetSpecialization(vlm_group_id="C64854.LB", bc_id=bc.bc_id, domain="LB", short_name="Ketones LB", status="published")
            db.session.add(spec)
            db.session.commit()
            wb = openpyxl.load_workbook(export_governance_xlsx([], [spec]))
        ws = wb["SDTM_LB"]
        assert ws.max_row == 2
        headers = [c.value for c in ws[1]]
        assert ws.cell(row=2, column=headers.index("vlm_group_id") + 1).value == "C64854.LB"

    def test_export_round_trips_through_ingestion_parser(self, app):
        with app.app_context():
            bc = self._make_bc()
            spec = DatasetSpecialization(vlm_group_id="C64854.LB", bc_id=bc.bc_id, domain="LB", short_name="Ketones LB", status="published")
            spec.variables = [{"sdtm_variable": "LBTESTCD", "data_type": "string", "mandatory_variable": "Y"}]
            db.session.add(spec)
            db.session.commit()
            buf = export_governance_xlsx([], [spec])

        records = parse_xlsx(buf)
        spec_records = [r for r in records if r.get("record_type") == "specialization"]
        assert len(spec_records) == 1
        rec = spec_records[0]
        assert rec["mapped"]["vlm_group_id"] == "C64854.LB"
        assert rec["mapped"]["bc_id"] == "C64854"
        assert rec["mapped"]["domain"] == "LB"
        assert rec["variables"][0]["sdtm_variable"] == "LBTESTCD"
        assert rec["variables"][0]["mandatory_variable"] == "Y"
        assert rec["errors"] == []


class TestExportOdmXml:
    def test_valid_odm_structure(self):
        xml = export_odm_xml(SAMPLE_BCS)
        root = etree.fromstring(xml.encode())
        ns = {"odm": "http://www.cdisc.org/ns/odm/v1.3"}
        assert root.tag == "{http://www.cdisc.org/ns/odm/v1.3}ODM"
        assert root.get("FileType") == "Snapshot"
        item_defs = root.findall("odm:ItemDef", ns)
        assert [i.get("OID") for i in item_defs] == ["C64849", "C25298"]
        # Definition text present
        text = item_defs[0].find("odm:Description/odm:TranslatedText", ns)
        assert text.text == "A quantitative measurement of HbA1c."

    def test_alias_only_when_ncit_code_present(self):
        xml = export_odm_xml(SAMPLE_BCS)
        root = etree.fromstring(xml.encode())
        ns = {"odm": "http://www.cdisc.org/ns/odm/v1.3"}
        item_defs = root.findall("odm:ItemDef", ns)
        assert item_defs[0].find("odm:Alias", ns) is not None
        assert item_defs[0].find("odm:Alias", ns).get("Name") == "C64849"
        assert item_defs[1].find("odm:Alias", ns) is None

    def test_empty_list(self):
        root = etree.fromstring(export_odm_xml([]).encode())
        assert len(root) == 0

    def test_decs_emit_item_defs(self):
        bcs_with_dec = [
            {
                **SAMPLE_BCS[0],
                "decs": [
                    {"dec_id": "C64849.DEC.1", "ncit_dec_code": "C999", "dec_label": "Result Value", "data_type": "decimal", "example_set": "", "required": False},
                ],
            }
        ]
        xml = export_odm_xml(bcs_with_dec)
        root = etree.fromstring(xml.encode())
        ns = {"odm": "http://www.cdisc.org/ns/odm/v1.3"}
        item_defs = root.findall("odm:ItemDef", ns)
        assert len(item_defs) == 2
        dec_item = next(i for i in item_defs if i.get("OID") == "C64849.DEC.1")
        assert dec_item.get("Name") == "Result Value"
        assert dec_item.get("DataType") == "float"
        alias = dec_item.find("odm:Alias", ns)
        assert alias.get("Name") == "C999"

    def test_bc_without_decs_key_emits_no_extra_item_defs(self):
        xml = export_odm_xml(SAMPLE_BCS)
        root = etree.fromstring(xml.encode())
        ns = {"odm": "http://www.cdisc.org/ns/odm/v1.3"}
        item_defs = root.findall("odm:ItemDef", ns)
        assert len(item_defs) == 2
